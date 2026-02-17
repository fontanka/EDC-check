
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import os
import re
import logging
import datetime
from io import BytesIO
import zipfile

from base_exporter import BaseExporter

logger = logging.getLogger(__name__)

# --- Configuration Constants (Ported from EDC Extractor) ---

# Mapping of Template Columns (Letter) to Semantic Key
TEMPLATE_SEMANTICS = {
  "G": "tr_color_grade", "H": "tr_hepatic_grade", "I": "tr_vcw_grade", "J": "tr_eroa_grade",
  "K": "tv_morphology", "L": "flail", "M": "color_flow_tr_jet", "N": "jet_area",
  "O": "eroa_numeric", "P": "regurg_vol", "Q": "rv_basal_diam", "R": "rv_mid_diam",
  "S": "rv_long_diam", "T": "rv_eda", "U": "rv_esa", "V": "rvfac", "W": "rv_eject_time",
  "X": "lv_syst_diam", "Y": "lv_diast_diam", "Z": "ra_area", "AA": "ra_diam",
  "AB": "lvef", "AC": "rvef", "AD": "tapse", "AE": "s_wave", "AF": "tcot",
  "AG": "mpi", "AH": "ivc_diam", "AI": "ivc_variation", "AJ": "delta_psys",
  "AK": "lvot_vti", "AL": "lvot_diam", "AM": "cardiac_output",
}

# Semantic Key -> Label Patterns (to find columns in DF)
SEMANTIC_PATTERNS = {
  "tr_hepatic_grade": ["TR evaluated based on systolic hepatic backflow"],
  "tr_color_grade": ["TR evaluated based on color doppler on the Trillium"],
  "tr_vcw_grade": ["Vena Contracta Width"],
  "tr_eroa_grade": ["TR evaluated based on EROA"],
  "tv_morphology": ["Tricuspid valve morphology"],
  "flail": ["Flail:"],
  "color_flow_tr_jet": ["Color flow TR jet"],
  "jet_area": ["Jet area, cm2"],
  "eroa_numeric": ["EROA, cm2"],
  "regurg_vol": ["regurgitant volume, ml"],
  "rv_basal_diam": ["RV basal diameter"],
  "rv_mid_diam": ["RV mid diameter"],
  "rv_long_diam": ["RV longitudinal diameter"],
  "rv_eda": ["RV EDA"],
  "rv_esa": ["RV ESA"],
  "rvfac": ["RVFAC", "RV FAC", "RV Fractional Area Change"],
  "rv_eject_time": ["RV Ejection Time"],
  "lv_syst_diam": ["LV systolic diameter"],
  "lv_diast_diam": ["LV diastolic diameter"],
  "ra_area": ["RA area"],
  "ra_diam": ["RA diameter"],
  "lvef": ["LVEF"],
  "rvef": ["RVEF"],
  "tapse": ["Tricuspid annular plane systolic excursion", "TAPSE"],
  "s_wave": ["Pulsed Doppler S wave"],
  "tcot": ["Closure to Opening Time", "TCOT"],
  "mpi": ["Pulsed Doppler MPI"],
  "ivc_diam": ["IVC diameter"],
  "ivc_variation": ["IVC respiratory variations"],
  "delta_psys": ["ΔPsys on Tricuspid Valve"],
  "lvot_vti": ["LVOT VTI", "LVOT VTI (cm)"],
  "lvot_diam": ["LVOT diameter", "LVOT diameter (cm)"],
  "cardiac_output": ["Cardiac Output"],
}

# Visit Name -> Data Prefix Logic (Main Sheet)
# Maps the visit name to the expected column prefixes in the main dataset
# Using _SP suffix columns for actual Sponsor (Core Lab) values
VISIT_CONFIG = {
    "Screening": {
        "prefix": "SBV_ECHO_SPONSOR_", 
        "date_col": "SBV_ECHO_SPONSOR_EGDTC",
        "type_col": "SBV_ECHO_SPONSOR_PRSCAT_ECHO"
    },
    "1-day pre-procedure": {
        "prefix": "TV_ECHO_1DPP_SPONSOR_", 
        "date_col": "TV_ECHO_1DPP_SPONSOR_EGDTC",
        "type_col": "TV_ECHO_1DPP_SPONSOR_PRSCAT_ECHO"
    },
    "Pre-procedure": {
        "prefix": "TV_ECHO_PRE_POST_SPONSOR_", 
        "suffix_filter": "_PRE", 
        "date_col": "TV_ECHO_PRE_POST_SPONSOR_EGDTC_PRE",
        "type_col": "TV_ECHO_PRE_POST_SPONSOR_PRSCAT_PRE_ECHO"
    },
    "Post-procedure": {
        "prefix": "TV_ECHO_PRE_POST_SPONSOR_", 
        "suffix_filter": "_POST", 
        "date_col": "TV_ECHO_PRE_POST_SPONSOR_EGDTC_POST_1",
        "type_col": "TV_ECHO_PRE_POST_SPONSOR_PRSCAT_POST_ECHO1"
    },
    "1-day post-procedure": {
        "prefix": "TV_ECHO_1D_SPONSOR_", 
        "date_col": "TV_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "TV_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "Discharge": {
        "prefix": "DV_ECHO_1D_SPONSOR_", 
        "date_col": "DV_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "DV_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "30-day": {
        "prefix": "FU1M_ECHO_1D_SPONSOR_", 
        "date_col": "FU1M_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "FU1M_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "6M": {
        "prefix": "FU6M_ECHO_1D_SPONSOR_", 
        "date_col": "FU6M_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "FU6M_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "1Y": {
        "prefix": "FU1Y_ECHO_1D_SPONSOR_", 
        "date_col": "FU1Y_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "FU1Y_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "2Y": {
        "prefix": "FU2Y_ECHO_1D_SPONSOR_", 
        "date_col": "FU2Y_ECHO_1D_SPONSOR_EGDTC",
        "type_col": "FU2Y_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
    "4Y": {
        "prefix": "FU4Y_ECHO_1D_SPONSOR_", 
        "date_col": "FU4Y_ECHO_1D_SPONSOR_EGDTC", 
        "type_col": "FU4Y_ECHO_1D_SPONSOR_PRSCAT_ECHO"
    },
}

VISIT_ORDER = [
  "Screening", "1-day pre-procedure", "Pre-procedure", "Post-procedure", 
  "1-day post-procedure", "Discharge", "30-day", "6M", "1Y", "2Y", "4Y"
]

class EchoExporter(BaseExporter):
    def __init__(self, df_main, template_path, labels_map):
        super().__init__(df_main, template_path, labels_map)
        self.col_map = {}

    def normalize_label(self, s):
        if not isinstance(s, str): return str(s)
        return re.sub(r'[^a-z0-9]', '', s.lower())

    def is_valid_value(self, val):
        """Check if value is valid (not NaN, not empty string, not 'nan' string)."""
        return self.is_valid(val)
        return True

    def find_column(self, visit_name, semantic_key):
        """Find column for a visit and semantic key, prioritizing _SP suffix columns."""
        cache_key = (visit_name, semantic_key)
        if cache_key in self.col_map:
            return self.col_map[cache_key]

        config = VISIT_CONFIG.get(visit_name)
        if not config: return None
        
        prefix = config['prefix']
        suffix_filter = config.get('suffix_filter')
        patterns = SEMANTIC_PATTERNS.get(semantic_key, [])

        # First pass: look for _SP suffix columns (actual Sponsor/Core Lab values)
        candidates = []
        for col in self.df_main.columns:
            if not str(col).startswith(prefix):
                continue
            
            # Handle suffix_filter for PRE/POST distinction
            # Column patterns: _FAORRES_PRE_ECHO1_SP vs _FAORRES_POST_ECHO1_SP
            if suffix_filter:
                # For _PRE: column should contain FAORRES_PRE_ (not FAORRES_POST_)
                # For _POST: column should contain FAORRES_POST_
                if suffix_filter == "_PRE":
                    if "FAORRES_PRE_" not in col and "_PRE_ECHO" not in col:
                        continue
                    if "FAORRES_POST_" in col or "_POST_ECHO" in col:
                        continue
                elif suffix_filter == "_POST":
                    if "FAORRES_POST_" not in col and "_POST_ECHO" not in col:
                        continue
                else:
                    if suffix_filter not in col:
                        continue
            
            # Label Check
            label = str(self.labels_map.get(col, col)).lower()
            
            # Check for patterns
            for p in patterns:
                if p.lower() in label:
                    candidates.append(col)
                    break
        
        # Prioritize _SP suffix columns (actual sponsor values)
        sp_cols = [c for c in candidates if c.endswith('_SP')]
        if sp_cols:
            self.col_map[cache_key] = sp_cols[0]
            return sp_cols[0]
        
        # Fallback to non-_SP columns
        if candidates:
            self.col_map[cache_key] = candidates[0]
            return candidates[0]
            
        return None

    def get_value(self, row, col_name):
        if not col_name or col_name not in row:
            return None
        val = row[col_name]
        if pd.isna(val) or str(val).strip() == "":
             return None
        return str(val)

    def check_visit_has_data(self, row, visit):
        """Check if a visit has actual data.
        First checks date columns (with and without SPONSOR prefix).
        If no date found, checks if any data columns (FAORRES) have values.
        """
        conf = VISIT_CONFIG.get(visit)
        if not conf:
            return False
        date_col = conf.get('date_col')
        prefix = conf.get('prefix', '')
        
        # Try the configured date column first
        if date_col and date_col in row.index:
            d_val = row[date_col]
            if pd.notna(d_val) and str(d_val).strip():
                return True
        
        # If configured column has no data, try alternative patterns
        # Sometimes data is in SBV_ECHO_EGDTC instead of SBV_ECHO_SPONSOR_EGDTC
        if date_col:
            # Try removing _SPONSOR from column name
            alt_col = date_col.replace('_SPONSOR', '')
            if alt_col in row.index:
                d_val = row[alt_col]
                if pd.notna(d_val) and str(d_val).strip():
                    return True
            
            # Try adding _SPONSOR if not present
            if '_SPONSOR' not in date_col:
                parts = date_col.split('_')
                # Insert SPONSOR before last part
                alt_col = '_'.join(parts[:-1]) + '_SPONSOR_' + parts[-1]
                if alt_col in row.index:
                    d_val = row[alt_col]
                    if pd.notna(d_val) and str(d_val).strip():
                        return True
        
        # Fallback: check if any FAORRES data columns have values for this visit
        # This catches cases where Core Lab has data but no date was entered
        if prefix:
            for col_name in row.index:
                if prefix in str(col_name) and 'FAORRES' in str(col_name):
                    val = row[col_name]
                    if pd.notna(val) and str(val).strip() and str(val).lower() != 'nan':
                        return True
        
        return False

    def get_visits_with_data(self, patient_id):
        """Return list of visits that have data for a patient."""
        rows = self.df_main[self.df_main['Screening #'] == patient_id]
        if rows.empty:
            return []
        row = rows.iloc[0]
        return [v for v in VISIT_ORDER if self.check_visit_has_data(row, v)]

    def process_patient(self, patient_id, selected_visits, delete_empty_rows=True):
        rows = self.df_main[self.df_main['Screening #'] == patient_id]
        if rows.empty:
            return None
        row = rows.iloc[0]

        try:
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active 
        except Exception as e:
            logger.error("Error loading template: %s", e)
            return None

        # IMMEDIATELY unmerge DATA cells in columns A and B (row 4+) to prevent MergedCell errors
        # Template has merged cells A4:A14 and B4:B14 that cause issues when writing/deleting
        # But KEEP header merges A2:A3 and B2:B3 intact
        merged_to_unmerge = []
        for merge_range in list(ws.merged_cells.ranges):
            # Only unmerge if in columns A or B AND starts at row 4 or later (data rows)
            if merge_range.min_col <= 2 and merge_range.min_row >= 4:
                merged_to_unmerge.append(str(merge_range))
        for merge_range in merged_to_unmerge:
            try:
                ws.unmerge_cells(merge_range)
            except ValueError:
                pass

        # Map template rows to visits
        visit_rows = {}
        # Sort visits by length (longest first) to avoid partial matches
        # e.g., "1-day post-procedure" should match before "Post-procedure"
        sorted_visits = sorted(VISIT_ORDER, key=lambda x: len(x), reverse=True)
        
        for r in range(1, 100): # Scan first 100 rows
            cell_val = ws.cell(row=r, column=4).value 
            if cell_val:
                norm_val = self.normalize_label(str(cell_val))
                for v_name in sorted_visits:
                    # Skip if already matched
                    if v_name in visit_rows:
                        continue
                    v_norm = self.normalize_label(v_name)
                    # Check for specific match patterns
                    if v_name == '2Y' and '2yfoloowup' in norm_val:
                        visit_rows[v_name] = r
                        break
                    elif v_name == '1-day post-procedure' and '1daypost' in norm_val:
                        visit_rows[v_name] = r
                        break
                    elif v_name == '1-day pre-procedure' and '1daypre' in norm_val:
                        visit_rows[v_name] = r
                        break
                    elif v_norm in norm_val:
                        visit_rows[v_name] = r
                        break
        
        # Track which visits have data and which rows to delete
        visits_with_data = set()
        rows_to_delete = []
        
        # Fill Data
        has_data = False
        for visit in selected_visits:
            if visit not in visit_rows:
                continue
            
            row_idx = visit_rows[visit]
            conf = VISIT_CONFIG.get(visit)
            
            # Check if visit has date (i.e., visit happened)
            visit_has_data = self.check_visit_has_data(row, visit)
            
            if not visit_has_data:
                # Mark row for deletion if visit didn't happen
                if delete_empty_rows:
                    rows_to_delete.append(row_idx)
                continue
            
            visits_with_data.add(visit)
            
            # 1. Date (format: dd-mmm-yyyy)
            date_col = conf.get('date_col')
            visit_date_obj = None
            
            if date_col and date_col in row:
                d_val = row[date_col]
                if pd.notna(d_val):
                    try:
                        visit_date_obj = pd.to_datetime(d_val)
                        # Format as dd-mmm-yyyy (e.g., 15-Aug-2025)
                        date_str = visit_date_obj.strftime('%d-%b-%Y')
                        ws.cell(row=row_idx, column=5).value = date_str
                    except (ValueError, TypeError):
                        # Fallback to simple date string
                        date_str = str(d_val).split('T')[0]
                        ws.cell(row=row_idx, column=5).value = date_str
            
            # 2. Age at Visit (Col C) - Calculate from DOB if available
            dob = None
            dob_is_partial = False
            
            # Try to get full date of birth
            if 'SBV_DM_BRTHDAT' in row.index and self.is_valid_value(row['SBV_DM_BRTHDAT']):
                try:
                    dob = pd.to_datetime(row['SBV_DM_BRTHDAT'])
                except (ValueError, TypeError):
                    pass
            
            # Check if DOB is partial (year only)
            if 'SBV_DM_BRTHDAT_PARTIAL' in row.index and self.is_valid_value(row['SBV_DM_BRTHDAT_PARTIAL']):
                dob_is_partial = True
            
            # If no full DOB but we have year, create partial DOB
            if dob is None and 'SBV_DM_BRTHDAT_YEAR' in row.index and self.is_valid_value(row['SBV_DM_BRTHDAT_YEAR']):
                try:
                    birth_year = int(float(str(row['SBV_DM_BRTHDAT_YEAR'])))
                    # Use July 1 as approximate mid-year
                    dob = pd.Timestamp(year=birth_year, month=7, day=1)
                    dob_is_partial = True
                except (ValueError, TypeError):
                    pass
            
            # Calculate age at visit from DOB
            if dob is not None and visit_date_obj is not None:
                try:
                    age_at_visit = int((visit_date_obj - dob).days / 365.25)
                    age_cell = ws.cell(row=row_idx, column=3)
                    age_cell.value = age_at_visit
                    
                    # If DOB is partial (year only), mark with bold italic
                    if dob_is_partial:
                        from openpyxl.styles import Font
                        age_cell.font = Font(bold=True, italic=True)
                except (ValueError, TypeError):
                    pass  # Skip age calculation if dates are invalid
            elif visit == "Screening":
                # Fallback: use DM_AGE for screening
                screening_age = None
                if 'SBV_DM_AGE' in row.index and self.is_valid_value(row['SBV_DM_AGE']):
                    screening_age = row['SBV_DM_AGE']
                elif 'DM_AGE' in row.index and self.is_valid_value(row['DM_AGE']):
                    screening_age = row['DM_AGE']
                if screening_age:
                    ws.cell(row=row_idx, column=3).value = screening_age

            # 3. Echo Type (Col F)
            type_col = conf.get('type_col')
            if type_col and type_col in row:
                t_val = str(row[type_col]).upper()
                if 'TTE' in t_val: ws.cell(row=row_idx, column=6).value = "TTE"
                elif 'TEE' in t_val: ws.cell(row=row_idx, column=6).value = "TEE"

            # 4. Metrics (Col G to AL)
            for col_letter, semantic_key in TEMPLATE_SEMANTICS.items():
                db_col = self.find_column(visit, semantic_key)
                col_idx = openpyxl.utils.cell.column_index_from_string(col_letter)
                
                if db_col:
                    val = self.get_value(row, db_col)
                    if val:
                        has_data = True
                        
                        # Special handling for color_flow_tr_jet - merge with specify columns
                        if semantic_key == "color_flow_tr_jet":
                            # Extract ECHO number from column name (e.g., ECHO5, ECHO6, ECHO3)
                            import re
                            echo_match = re.search(r'ECHO(\d+)', db_col)
                            echo_num = echo_match.group(1) if echo_match else '6'
                            
                            merged_parts = [val]
                            
                            # Find base column without _SP suffix
                            base_col = db_col[:-3] if db_col.endswith('_SP') else db_col
                            
                            # Try to get additional values (e.g., ECHO6_1, ECHO6_2 or ECHO5_1, ECHO5_2)
                            for suffix in ['_1', '_2']:
                                extra_col = base_col.replace(f'ECHO{echo_num}', f'ECHO{echo_num}{suffix}')
                                if db_col.endswith('_SP'):
                                    extra_col = extra_col + '_SP'
                                extra_val = self.get_value(row, extra_col)
                                # Only add if different from base value and valid
                                if extra_val and extra_val.lower() not in ['nan', 'none', '']:
                                    # Avoid adding duplicate values
                                    if extra_val.lower() != val.lower() and extra_val not in merged_parts:
                                        merged_parts.append(extra_val)
                            
                            merged_val = ". ".join(merged_parts)
                            ws.cell(row=row_idx, column=col_idx).value = merged_val
                        else:
                            ws.cell(row=row_idx, column=col_idx).value = val
                        
                        # Apply text wrapping with vertical alignment
                        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='center')
                    else:
                        # Column exists but no value - check if status column indicates "Not done"
                        # Try to find corresponding FASTAT column (e.g., FAORRES_ECHO6 -> FASTAT_ECHO6)
                        # Try multiple column name variations
                        stat_val = None
                        stat_col_base = db_col.replace('FAORRES', 'FASTAT')
                        
                        # Try different variations: with _SP, without _SP, base name
                        stat_col_variations = [
                            stat_col_base,  # Same as FAORRES column (with _SP if original had it)
                            stat_col_base[:-3] if stat_col_base.endswith('_SP') else stat_col_base + '_SP',  # Toggle _SP
                        ]
                        
                        for stat_col in stat_col_variations:
                            stat_val = self.get_value(row, stat_col)
                            if stat_val and str(stat_val).lower() not in ['nan', 'none', '']:
                                break
                        
                        # FASTAT = "true" means the "not done" checkbox was checked
                        if stat_val and (str(stat_val).lower() == 'true' or 'not done' in str(stat_val).lower()):
                            ws.cell(row=row_idx, column=col_idx).value = "Not done"
                            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='center')
                        # else: leave cell blank (missing data)
            
            # Auto-adjust row height for wrapped text (estimate based on content)
            max_lines = 1
            for col in range(1, 39):  # A to AL
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val:
                    # Estimate lines based on length and typical column width
                    text_len = len(str(cell_val))
                    estimated_lines = max(1, (text_len // 15) + 1)  # Rough estimate
                    max_lines = max(max_lines, estimated_lines)
            
            # Set row height (default is ~15, increase for multi-line content)
            if max_lines > 1:
                ws.row_dimensions[row_idx].height = max(15, min(max_lines * 14, 60))
        
        # Delete rows for visits without data (in reverse order to preserve row indices)
        if delete_empty_rows and rows_to_delete:
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
        
        # Also delete rows for visits not selected (clean up template)
        if delete_empty_rows:
            # Re-scan to find remaining visit rows that weren't selected
            remaining_rows_to_delete = []
            for r in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=4).value
                if cell_val:
                    norm_val = self.normalize_label(str(cell_val))
                    for v_name in VISIT_ORDER:
                        v_norm = self.normalize_label(v_name)
                        if v_norm in norm_val or (v_name == '2Y' and '2yfoloowup' in norm_val):
                            if v_name not in selected_visits or v_name not in visits_with_data:
                                remaining_rows_to_delete.append(r)
                            break
            
            for row_idx in sorted(remaining_rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
        
        # Fix borders for merged cells (A4 and B4) after row deletion
        # Apply bottom border to the last row that has data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find the last data row (after deletions)
        last_data_row = 4  # Default to row 4 if only header
        for r in range(ws.max_row, 3, -1):
            if ws.cell(row=r, column=4).value:  # Visit column D has data
                last_data_row = r
                break
        
        # Set Header Info (cells already unmerged at start of process_patient)
        ws['A4'] = str(patient_id)
        # Gender - check multiple possible columns
        gender = None
        for g_col in ['SBV_DM_GENDER', 'DM_GENDER', 'DM_SEX', 'SBV_DM_SEX']:
            if g_col in row.index:
                g_val = row[g_col]
                if pd.notna(g_val) and str(g_val).strip():
                    gender = str(g_val).strip().capitalize()
                    break
        if gender:
            ws['B4'] = gender
        
        # Apply bottom border to ALL cells in the last row (now safe - A and B unmerged)
        for col in range(1, 40):  # A to AM (extended range)
            cell = ws.cell(row=last_data_row, column=col)
            # Apply complete border with medium bottom for consistency
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='medium')  # Medium bottom border like template
            )
        
        # Now re-merge A4 to A[last_data_row] and B4 to B[last_data_row]
        if last_data_row > 4:
            try:
                ws.merge_cells(f'A4:A{last_data_row}')
                ws.merge_cells(f'B4:B{last_data_row}')
            except ValueError:
                pass  # Cells might already be merged or have issues

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def generate_export(self, patient_ids, selected_visits, delete_empty_rows=True):
        """Generates export - single xlsx for one patient, ZIP for multiple patients."""
        return super().generate_export(
            patient_ids,
            selected_visits=selected_visits,
            delete_empty_rows=delete_empty_rows,
        )
