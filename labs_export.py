
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime
from io import BytesIO
import zipfile

# --- Configuration Constants ---

# Template row mapping: row number (1-indexed) -> (parameter_name, test_code, lab_type)
# Fixed: CREA not CREAT, Troponin/NT-proBNP are in ENZ not BM
TEMPLATE_ROW_MAP = {
    4: ("Blood urea nitrogen (BUN)", "BUN", "BMP"),
    5: ("Glucose", "GLUC", "BMP"),
    6: ("Creatinine", "CREA", "BMP"),  # Fixed: CREA not CREAT
    7: ("eGFR", "eGFR", "BMP"),  # Calculated field - special handling
    8: ("Sodium", "SODIUM", "BMP"),
    9: ("Serum Chloride", "CL", "BMP"),
    10: ("Potassium", "K", "BMP"),
    11: ("Bilirubin Total", "BILI", "LFP"),
    12: ("AST (GOT)", "AST", "LFP"),
    13: ("ALT (GPT)", "ALT", "LFP"),
    14: ("GGT", "GGT", "LFP"),
    15: ("LDH", "LDH", "LFP"),  # Fixed: LDH is in LFP not ENZ
    16: ("PT-INR", "PTI", "COA"),
    17: ("PT", "PT", "COA"),
    18: ("APTT-sec", "APTT", "COA"),
    19: ("Troponin T/I", ["TROPONT", "TROPONI"], "ENZ"),  # Check both, label which one
    20: ("NT-proBNP", "BNPPRO", "ENZ"),  # Fixed: ENZ not BM
    21: ("RBC", "RBC", "CBC"),
    22: ("HB", "HGB", "CBC"),
    23: ("HCT", "HCT", "CBC"),
    24: ("PLT", "PLAT", "CBC"),
    25: ("WBC", "WBC", "CBC"),
    26: ("MCV", "MCV", "CBC"),
    27: ("MCH", "MCH", "CBC"),
    28: ("MCHC", "MCHC", "CBC"),
    29: ("RDW", "RDW", "CBC"),
    30: ("NEUT%", "NEUTP", "CBC"),
    31: ("NEUT abs", "NEUTA", "CBC"),
    32: ("LYM%", "LYMPP", "CBC"),
    33: ("LYMP abs", "LYMPA", "CBC"),
    34: ("MONO%", "MONOP", "CBC"),
    35: ("MONO abs", "MONOA", "CBC"),
    36: ("EOS%", "EOSP", "CBC"),
    37: ("EOS abs", "EOSA", "CBC"),
    38: ("BASO%", "BASOP", "CBC"),
    39: ("BASO abs", "BASOA", "CBC"),
    40: ("Procalcitonin", "PCT", "BM"),
    41: ("C-Reactive Protein (CRP)", "CRP", "BM"),
}

# Visit configuration for non-daily visits
LABS_VISIT_CONFIG = {
    "Screening": {"prefix": "SBV_LB_", "date_col": "SBV_SV_SVSTDTC"},
    "Discharge": {"prefix": "DV_LB_", "date_col": "DV_SV_SVSTDTC"},  # Separate discharge visit
    "30D": {"prefix": "FU1M_LB_", "date_col": "FU1M_SV_SVSTDTC"},
    "6M": {"prefix": "FU6M_LB_", "date_col": "FU6M_SV_SVSTDTC"},
    "1Y": {"prefix": "FU1Y_LB_", "date_col": "FU1Y_SV_SVSTDTC"},
    "2Y": {"prefix": "FU2Y_LB_", "date_col": "FU2Y_SV_SVSTDTC"},
    "4Y": {"prefix": "FU4Y_LB_", "date_col": "FU4Y_SV_SVSTDTC"},
}

# Visit order for follow-up visits (after discharge)
FOLLOWUP_VISITS = ["30D", "6M", "1Y", "2Y", "4Y"]


# Unit Conversion Factors
# Factors are multiplier to convert FROM key1 TO key2
# keys are normalization(source_unit), normalization(target_unit)
CONVERSION_FACTORS = {
    'CRP': {
        ('mg/dl', 'mg/l'): 10.0,
        ('mg/l', 'mg/dl'): 0.1,
    }
}

# Unconvertible unit pairs - these cannot be mathematically converted
# When both are present, display both with color coding
UNCONVERTIBLE_PAIRS = [
    frozenset({'fl', '%'}),  # RDW-SD (fL) vs RDW-CV (%) - genuinely different measurements
]

# Colors for dual-unit display (hex without #, for openpyxl)
DUAL_UNIT_COLORS = {
    'unit1': '0000FF',  # Blue
    'unit2': '008000',  # Green
}

# Color for out-of-range values
OUT_OF_RANGE_COLOR = 'FF0000'  # Red

class LabsExporter:
    def __init__(self, df_main, template_path, labels_map, unit_callback=None, highlight_out_of_range=False):
        self.df_main = df_main
        self.template_path = template_path
        self.labels_map = labels_map
        self.col_cache = {}
        self.unit_callback = unit_callback
        self.highlight_out_of_range = highlight_out_of_range
        self.param_target_units = {} # param_code -> target_unit
    
    def normalize_unit(self, unit):
        """Normalize unit string for comparison."""
        if not unit:
            return None
        
        s = str(unit).strip()
        
        # Remove pipe artifacts (e.g., "||||mg/L|||" -> "mg/L", "%|%" -> "%")
        s = s.strip('|')
        # If still contains only pipes or is empty after strip
        if not s or s == '|':
            return None
        # Handle internal pipes (take first non-empty part)
        if '|' in s:
            parts = [p.strip() for p in s.split('|') if p.strip()]
            s = parts[0] if parts else None
            if not s:
                return None
        
        # Lowercase for case-insensitive comparison
        s = s.lower()
        
        # Treat placeholder values as "no unit" (dimensionless or missing)
        if s in ['-', 'n/a', 'na', 'none', 'nan', '']:
            return None
        
        # Normalize equivalent units
        equivalents = {
            'secs': 'sec',
            'seconds': 'sec',
            'iu/l': 'u/l',
            'ratio': None,  # dimensionless, treat as no unit
        }
        if s in equivalents:
            return equivalents[s]
        
        return s

    def convert_value(self, val, source_unit, target_unit, param_code):
        """Convert value from source_unit to target_unit."""
        if val is None or not source_unit or not target_unit:
            return val
            
        s_unit = self.normalize_unit(source_unit)
        t_unit = self.normalize_unit(target_unit)
        
        if s_unit == t_unit:
            return val
            
        num_val = self.to_number(val)
        if num_val is None or isinstance(num_val, str):
            return val
            
        # Look up conversion factor
        factors = CONVERSION_FACTORS.get(param_code, {})
        factor = factors.get((s_unit, t_unit))
        
        # Generic conversions if no specific factor found
        if factor is None:
            if s_unit == 'mg/dl' and t_unit == 'mg/l':
                factor = 10.0
            elif s_unit == 'mg/l' and t_unit == 'mg/dl':
                factor = 0.1
            # ng/mL to ng/L (and reverse) - Troponin, etc.
            elif s_unit == 'ng/ml' and t_unit == 'ng/l':
                factor = 1000.0
            elif s_unit == 'ng/l' and t_unit == 'ng/ml':
                factor = 0.001
            # pg/mL to pg/L (and reverse) - NT-proBNP, etc.
            elif s_unit == 'pg/ml' and t_unit == 'pg/l':
                factor = 1000.0
            elif s_unit == 'pg/l' and t_unit == 'pg/ml':
                factor = 0.001
        
        if factor is not None:
            new_val = num_val * factor
            # Round to reasonable decimals (e.g. 2 for now, or match input precision?)
            # For CRP, 1 decimal is usually enough
            return round(new_val, 2)
            
        return val

    def scan_parameter_units(self, row, prefix_list, lab_type, test_code, num_daily_days):
        """Scan all visits for units of a parameter."""
        found_units = set()
        
        # Check standard prefixes (Screening, Follow-ups)
        for prefix in prefix_list:
             u = self.get_units(row, prefix, lab_type, test_code)
             normalized = self.normalize_unit(u)
             if normalized:
                 found_units.add(normalized)
        
        # Check Daily Labs (TV_LB_)
        # We must check each day as units *could* theoretically change
        for i in range(num_daily_days):
            u = self.get_units(row, "TV_LB_", lab_type, test_code, pipe_index=i)
            normalized = self.normalize_unit(u)
            if normalized:
                found_units.add(normalized)
                
        return found_units

    def resolve_units(self, row, num_daily_days, patient_id):
        """Resolve unit conflicts for all parameters.
        
        Returns dict: template_row -> unit_info
        unit_info can be:
          - None (no units found)
          - str (single unit)
          - dict {'dual': True, 'units': [unit1, unit2]} for unconvertible pairs
        """
        resolved_map = {}
        
        # Group prefixes to scan
        non_daily_prefixes = [cfg['prefix'] for cfg in LABS_VISIT_CONFIG.values()]
        
        for template_row, (param_name, test_code_raw, lab_type) in TEMPLATE_ROW_MAP.items():
            # Handle list of test codes (e.g. Troponin)
            test_codes = test_code_raw if isinstance(test_code_raw, list) else [test_code_raw]
            
            all_found = set()
            for tc in test_codes:
                found = self.scan_parameter_units(row, non_daily_prefixes, lab_type, tc, num_daily_days)
                all_found.update(found)
            
            # Clean up: remove None
            all_found = {u for u in all_found if u}
            
            if not all_found:
                resolved_map[template_row] = None
                continue
                
            if len(all_found) == 1:
                resolved_map[template_row] = list(all_found)[0]
            else:
                # Conflict! Check if it's an unconvertible pair
                is_unconvertible = False
                for pair in UNCONVERTIBLE_PAIRS:
                    if all_found == pair or all_found.issubset(pair):
                        is_unconvertible = True
                        break
                
                if is_unconvertible:
                    # Return dual-unit info (no user prompt needed)
                    units_list = sorted(list(all_found))
                    resolved_map[template_row] = {
                        'dual': True,
                        'units': units_list,
                        'param_name': param_name
                    }
                else:
                    if self.unit_callback:
                        target = self.unit_callback(param_name, sorted(list(all_found)), patient_id)
                        resolved_map[template_row] = target
                    else:
                        resolved_map[template_row] = sorted(list(all_found))[0]
        
        return resolved_map

    def is_valid_value(self, val):
        """Check if value is valid (not NaN, not empty string, not 'nan' string)."""
        if pd.isna(val):
            return False
        val_str = str(val).strip().lower()
        if val_str in ['', 'nan', 'none']:
            return False
        return True
    
    def is_outside_reference_range(self, value, ref_min, ref_max):
        """Check if a numeric value is outside the reference range.
        
        Returns True if value < ref_min or value > ref_max.
        Returns False if value is within range, or if ranges/value are missing.
        """
        if value is None or (ref_min is None and ref_max is None):
            return False
        
        # Convert value to number
        num_val = self.to_number(value)
        if num_val is None or isinstance(num_val, str):
            return False  # Can't compare non-numeric values
        
        # Check against min bound
        if ref_min is not None:
            try:
                min_num = float(ref_min)
                if num_val < min_num:
                    return True
            except (ValueError, TypeError):
                pass
        
        # Check against max bound
        if ref_max is not None:
            try:
                max_num = float(ref_max)
                if num_val > max_num:
                    return True
            except (ValueError, TypeError):
                pass
        
        return False
    
    def to_number(self, val):
        """Convert string value to number if possible, otherwise return as string."""
        if val is None:
            return None
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ['nan', 'none', 'not done']:
            return val_str if val_str else None
        try:
            # Try integer first
            if '.' not in val_str:
                return int(val_str)
            # Then float
            return float(val_str)
        except ValueError:
            return val_str
    
    def get_treatment_date(self, row):
        """Get treatment date from TV_PR_SVDTC column."""
        date_col = "TV_PR_SVDTC"
        if date_col in row.index:
            val = row[date_col]
            if self.is_valid_value(val):
                try:
                    return pd.to_datetime(str(val).split('T')[0])
                except:
                    pass
        return None
    
    def get_daily_lab_dates(self, row):
        """Get list of dates from daily lab columns."""
        for lab_type in ['BMP', 'CBC', 'LFP', 'ENZ', 'COA']:
            date_col = f"TV_LB_{lab_type}_DV_LBDAT_{lab_type}"
            if date_col in row.index:
                val = row[date_col]
                if self.is_valid_value(val):
                    return [d.strip() for d in str(val).split('|') if d.strip()]
        return []
    
    def calculate_day_offsets(self, row):
        """Calculate day offsets from treatment date for each daily lab date.
        
        Returns list of tuples: [(date_str, day_offset, formatted_date), ...]
        where day_offset is relative to treatment date (e.g., -3, +1, +2, etc.)
        """
        treatment_date = self.get_treatment_date(row)
        daily_dates = self.get_daily_lab_dates(row)
        
        if not treatment_date or not daily_dates:
            return []
        
        result = []
        for date_str in daily_dates:
            try:
                dt = pd.to_datetime(date_str.split('T')[0])
                day_offset = (dt - treatment_date).days
                formatted_date = dt.strftime('%d-%b-%Y')
                result.append((date_str, day_offset, formatted_date))
            except:
                result.append((date_str, None, date_str))
        
        return result
    
    def find_lab_column(self, prefix, lab_type, test_code, col_type="LBORRES"):
        """Find the column name for a specific lab test."""
        cache_key = (prefix, lab_type, test_code, col_type)
        if cache_key in self.col_cache:
            return self.col_cache[cache_key]
        
        # Build possible column name patterns
        patterns = []
        
        # For daily labs (TV_LB_*_DV_*)
        if prefix == "TV_LB_":
            patterns.append(f"TV_LB_{lab_type}_DV_{col_type}_{test_code}")
        
        # Standard pattern: {PREFIX}{LAB_TYPE}_{COL_TYPE}_{TEST}
        patterns.append(f"{prefix}{lab_type}_{col_type}_{test_code}")
        
        for pattern in patterns:
            if pattern in self.df_main.columns:
                self.col_cache[cache_key] = pattern
                return pattern
        
        # Fallback: search for matching column
        for col in self.df_main.columns:
            col_str = str(col)
            if col_str.startswith(prefix) and lab_type in col_str and test_code in col_str and col_type in col_str:
                self.col_cache[cache_key] = col
                return col
        
        return None
    
    def get_lab_value(self, row, prefix, lab_type, test_code, pipe_index=None):
        """Get lab result value, optionally at a specific pipe index."""
        col = self.find_lab_column(prefix, lab_type, test_code, "LBORRES")
        
        # Special handling for eGFR
        if not col and test_code == "eGFR":
            if prefix == "TV_LB_":
                col = f"TV_LB_{lab_type}_DV_eGFR"
            else:
                col = f"{prefix}{lab_type}_eGFR"
            if col not in self.df_main.columns:
                return None
        
        if not col or col not in row.index:
            return None
            
        val = row[col]
        if not self.is_valid_value(val):
            return None
        
        val_str = str(val)
        if pipe_index is not None and '|' in val_str:
            parts = val_str.split('|')
            if pipe_index < len(parts):
                part = parts[pipe_index].strip()
                return part if part and part.lower() not in ['nan', 'none', ''] else None
            return None
        
        return val_str.strip()
    
    def get_troponin_value(self, row, prefix, pipe_index=None):
        """Get Troponin value with type label (T or I)."""
        lab_type = "ENZ"
        
        # Check Troponin T first
        val_t = self.get_lab_value(row, prefix, lab_type, "TROPONT", pipe_index)
        if val_t:
            return f"{val_t} (T)"
        
        # Check Troponin I
        val_i = self.get_lab_value(row, prefix, lab_type, "TROPONI", pipe_index)
        if val_i:
            return f"{val_i} (I)"
        
        return None
    
    def update_sheet1_columns(self, wb, num_daily_days, day_data, patient_ref_values=None):
        """Update Sheet1 reference range columns to match daily columns count.
        
        Sheet1 structure:
        - Row 4: Day labels (-1, 1, 2, 3, 4...)
        - Rows 6-7: WBC min/max
        - Rows 9-10: Cr min/max
        - Rows 12-13: PLT min/max
        - Rows 15-16: CRP min/max
        
        Columns B-F in template correspond to 5 days (baseline through day+4).
        We need to adjust to match num_daily_days and populate with patient's actual ref values.
        """
        if 'Sheet1' not in wb.sheetnames:
            return
        
        ws1 = wb['Sheet1']
        template_cols = 5  # B through F in template
        
        # Reference rows mapping: Sheet1 row -> (min_row, max_row)
        ref_rows = {
            'WBC': (6, 7),
            'Cr': (9, 10),
            'PLT': (12, 13),
            'CRP': (15, 16),
        }
        
        # Handle column count adjustment
        # Sheet1 needs columns for Daily days + Discharge
        total_chart_cols = num_daily_days + 1  # Daily + Discharge
        
        if total_chart_cols > template_cols:
            # Insert additional columns
            extra_cols = total_chart_cols - template_cols
            insert_at = 7  # Column G (after F)
            
            for _ in range(extra_cols):
                ws1.insert_cols(insert_at)
        
        elif total_chart_cols < template_cols:
            # Delete excess columns (from right to left)
            cols_to_delete = template_cols - total_chart_cols
            for _ in range(cols_to_delete):
                ws1.delete_cols(2 + total_chart_cols)
        
        # Update day labels in row 4 based on day_data
        for idx, (date_str, day_offset, formatted_date) in enumerate(day_data):
            col = 2 + idx  # Start at column B
            if day_offset is not None:
                ws1.cell(row=4, column=col).value = day_offset
        
        # Add "Discharge" label for the last column
        discharge_label_col = 2 + num_daily_days  # Column after last daily
        ws1.cell(row=4, column=discharge_label_col).value = "Disch"
        
        # Populate reference values with patient's actual values (constant across all columns including Discharge)
        if patient_ref_values:
            for label, (min_row, max_row) in ref_rows.items():
                ref_data = patient_ref_values.get(label, (None, None))
                if len(ref_data) == 3:
                     ref_min, ref_max, _ = ref_data
                else:
                     ref_min, ref_max = ref_data
                
                if ref_min is not None or ref_max is not None:
                    # Include all columns: daily + Discharge
                    for col_idx in range(total_chart_cols):
                        col = 2 + col_idx  # Start at column B
                        if ref_min is not None:
                            ws1.cell(row=min_row, column=col).value = self.to_number(ref_min)
                        if ref_max is not None:
                            ws1.cell(row=max_row, column=col).value = self.to_number(ref_max)
        
        # Update chart series formulas to match new column range
        # Check all worksheets for charts (not just Sheet2)
        last_col_letter = get_column_letter(1 + total_chart_cols)  # B + total_chart_cols - 1
        
        for sheet_name in wb.sheetnames:
            ws_check = wb[sheet_name]
            if hasattr(ws_check, '_charts') and ws_check._charts:
                for chart in ws_check._charts:
                    for series in chart.series:
                        if hasattr(series, 'val') and series.val and hasattr(series.val, 'numRef'):
                            current_ref = series.val.numRef.f
                            if current_ref and 'Sheet1' in current_ref:
                                # Update Sheet1 reference to match new column range
                                import re
                                # Match pattern like Sheet1!$B$6:$F$6
                                match = re.search(r"Sheet1!\$B\$(\d+):\$[A-Z]+\$(\d+)", current_ref)
                                if match:
                                    row_num = match.group(1)
                                    new_ref = f"Sheet1!$B${row_num}:${last_col_letter}${row_num}"
                                    series.val.numRef.f = new_ref
    
    def get_lab_status(self, row, prefix, lab_type, test_code, pipe_index=None):
        """Check if lab test was marked as 'not done'."""
        col = self.find_lab_column(prefix, lab_type, test_code, "LBSTAT")
        if not col or col not in row.index:
            return None
        
        val = row[col]
        if not self.is_valid_value(val):
            return None
        
        val_str = str(val)
        if pipe_index is not None and '|' in val_str:
            parts = val_str.split('|')
            if pipe_index < len(parts):
                val_str = parts[pipe_index].strip()
            else:
                return None
        
        val_lower = val_str.lower()
        if val_lower in ['true', 'yes', '1', 'not done', 'not performed']:
            return "Not done"
        return None
    
    def get_reference_values(self, row, lab_type, test_code):
        """Get reference min/max values and unit from Screening visit, with fallback to daily labs.
        
        Returns (ref_min, ref_max, ref_unit) tuple.
        """
        ref_min = None
        ref_max = None
        ref_unit = None
        
        # Try Screening visit first
        prefix = "SBV_LB_"
        min_col = self.find_lab_column(prefix, lab_type, test_code, "LBORNRLO")
        if min_col and min_col in row.index:
            val = row[min_col]
            if self.is_valid_value(val):
                ref_min = str(val).strip()
        
        max_col = self.find_lab_column(prefix, lab_type, test_code, "LBORNRHI")
        if max_col and max_col in row.index:
            val = row[max_col]
            if self.is_valid_value(val):
                ref_max = str(val).strip()
        
        # If we got values from screening, get the screening unit
        if ref_min is not None or ref_max is not None:
            ref_unit = self.get_units(row, prefix, lab_type, test_code)
        
        # Fallback to daily labs (TV_LB_) if screening values are missing
        if ref_min is None or ref_max is None:
            tv_prefix = "TV_LB_"
            used_tv_fallback = False
            
            if ref_min is None:
                tv_min_col = self.find_lab_column(tv_prefix, lab_type, test_code, "LBORNRLO")
                if tv_min_col and tv_min_col in row.index:
                    val = row[tv_min_col]
                    if self.is_valid_value(val):
                        # Take first value from pipe-delimited string
                        val_str = str(val).split('|')[0].strip()
                        if val_str and val_str.lower() not in ['nan', 'none', '']:
                            ref_min = val_str
                            used_tv_fallback = True
            
            if ref_max is None:
                tv_max_col = self.find_lab_column(tv_prefix, lab_type, test_code, "LBORNRHI")
                if tv_max_col and tv_max_col in row.index:
                    val = row[tv_max_col]
                    if self.is_valid_value(val):
                        # Take first value from pipe-delimited string
                        val_str = str(val).split('|')[0].strip()
                        if val_str and val_str.lower() not in ['nan', 'none', '']:
                            ref_max = val_str
                            used_tv_fallback = True
            
            # If we used TV fallback and don't have a unit yet, get from TV
            if used_tv_fallback and ref_unit is None:
                ref_unit = self.get_units(row, tv_prefix, lab_type, test_code, pipe_index=0)
        
        return ref_min, ref_max, ref_unit
    
    def get_units(self, row, prefix, lab_type, test_code, pipe_index=None):
        """Get units for a lab test."""
        col = self.find_lab_column(prefix, lab_type, test_code, "LBORRESU")
        if not col or col not in row.index:
            return None
        
        val = row[col]
        if not self.is_valid_value(val):
            return None
        
        val_str = str(val)
        if pipe_index is not None and '|' in val_str:
            parts = val_str.split('|')
            if pipe_index < len(parts):
                val_str = parts[pipe_index].strip()
            else:
                return None
        
        # Handle "Other" units
        if val_str.lower() == "other":
            oth_col = self.find_lab_column(prefix, lab_type, test_code, "LBORRESU_OTH")
            if oth_col and oth_col in row.index:
                oth_val = row[oth_col]
                if self.is_valid_value(oth_val):
                    return str(oth_val).strip()
        
        return val_str.strip()
    
    def format_date(self, date_str):
        """Format date string to dd-mmm-yyyy."""
        if not date_str:
            return ""
        try:
            if '|' in str(date_str):
                date_str = str(date_str).split('|')[0].strip()
            dt = pd.to_datetime(date_str)
            return dt.strftime('%d-%b-%Y')
        except:
            return str(date_str).split('T')[0] if 'T' in str(date_str) else str(date_str)
    
    def format_day_header(self, day_offset, is_discharge=False):
        """Format day column header based on offset from treatment."""
        if day_offset < 0:
            header = f"Baseline (Day {day_offset})"
        elif day_offset == 0:
            header = "Treatment Day"
        else:
            header = f"Day +{day_offset}"
        
        if is_discharge:
            header += " (Discharge)"
        
        return header
    
    def copy_column_format(self, ws, source_col, target_col, max_row=41):
        """Copy cell formatting from source column to target column."""
        from copy import copy
        for row_num in range(1, max_row + 1):
            source_cell = ws.cell(row=row_num, column=source_col)
            target_cell = ws.cell(row=row_num, column=target_col)
            
            # Copy styles
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
    
    def update_chart_labels(self, ws, patient_id, chart_ref_values):
        """Update chart titles with real reference ranges and series names with patient ID.
        
        Chart 0: Leukocyte (WBC) - rows 6-7 in Sheet1
        Chart 1: Creatinin (Cr) - rows 9-10 in Sheet1
        Chart 2: Platelet (PLT) - rows 12-13 in Sheet1
        Chart 3: CRP - rows 15-16 in Sheet1
        """
        
        if not hasattr(ws, '_charts') or not ws._charts:
            return
        
        # Map chart index to parameter info
        # Chart titles: "Parameter\nNormal values: min-max [units]"
        chart_info = [
            {'name': 'Leukocyte level', 'param': 'WBC', 'unit': 'K/micl'},
            {'name': 'Creatinin level', 'param': 'Cr', 'unit': 'mg/dL'},
            {'name': 'Platelet Count', 'param': 'PLT', 'unit': 'K/micl'},
            {'name': 'CRP level', 'param': 'CRP', 'unit': 'mg/dL'},
        ]
        
        for chart_idx, chart in enumerate(ws._charts):
            if chart_idx >= len(chart_info):
                break
            
            info = chart_info[chart_idx]
            # Unpack 3 values: min, max, unit (or default to None)
            ref_data = chart_ref_values.get(info['param'], (None, None, None))
            if len(ref_data) == 3:
                ref_min, ref_max, ref_unit = ref_data
            else:
                ref_min, ref_max = ref_data
                ref_unit = None
            
            # Use dynamic unit if available, otherwise fallback to default
            display_unit = ref_unit if ref_unit else info['unit']
            
            try:
                # Update chart title with real reference values
                # Set overlay=False to use "Above Chart" positioning instead of "Centered Overlay"
                if ref_min is not None and ref_max is not None:
                    # Current format: "Name (Normal: min-max unit)"
                    # Note: User's screenshot showed "Normal values:", which implies this code wasn't running
                    # or failing.
                    new_title = f"{info['name']} (Normal: {ref_min}-{ref_max} {display_unit})"
                    chart.title = new_title
                    # Set overlay to False
                    if hasattr(chart, 'title') and chart.title is not None:
                        chart.title.overlay = False
                
                # Update Y-Axis Title if present
                if hasattr(chart, 'y_axis') and chart.y_axis and chart.y_axis.title:
                    axis_title = f"{info['param']} [{display_unit}]"
                    chart.y_axis.title = axis_title
                    # Fix overlap like we did for main title
                    try:
                        chart.y_axis.title.overlay = False
                    except:
                        pass

                # Update first series name (RE -> patient_id) using SeriesLabel
                from openpyxl.chart.series import SeriesLabel
                if chart.series and len(chart.series) > 0:
                    chart.series[0].tx = SeriesLabel(v=str(patient_id))
                    
            except Exception as e:
                print(f"Error updating chart {chart_idx} labels: {e}")

    
    def update_charts(self, ws, num_daily_days, day_data):
        """Update chart data ranges for peri-procedural data including Discharge.
        
        Charts show daily data (columns 7 to 7+num_daily_days-1) PLUS Discharge column.
        Not Screening (col 6) or follow-up visits.
        
        Instead of rebuilding charts, we update the data references of existing series.
        """
        from openpyxl.chart import Reference
        from openpyxl.chart.series import SeriesLabel
        
        if not hasattr(ws, '_charts') or not ws._charts:
            return
        
        # Data range: columns 7 to discharge_col (inclusive)
        # Layout: G(7)=Day-1, ... | Discharge | 30D...
        min_col = 7  # Start at Baseline
        discharge_col = 7 + num_daily_days  # Discharge column is right after daily
        max_col = discharge_col  # Include Discharge in charts
        
        # Update each chart's series to use the correct column range
        for chart_idx, chart in enumerate(ws._charts):
            try:
                # Update each existing series to reference correct column range
                # ONLY update Sheet2 data series, NOT Sheet1 reference range series
                for series in chart.series:
                    if hasattr(series, 'val') and series.val and hasattr(series.val, 'numRef'):
                        current_ref = series.val.numRef.f
                        if current_ref:
                            # Skip Sheet1 references (these are constant ref range lines)
                            if 'Sheet1' in current_ref:
                                continue
                            
                            # Parse the row from Sheet2 reference
                            import re
                            match = re.search(r'\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)', current_ref)
                            if match:
                                data_row = int(match.group(2))
                                # Create new reference with updated column range
                                sheet_name = ws.title
                                min_col_letter = get_column_letter(min_col)
                                max_col_letter = get_column_letter(max_col)
                                new_ref = f"'{sheet_name}'!${min_col_letter}${data_row}:${max_col_letter}${data_row}"
                                series.val.numRef.f = new_ref
                
                # Update category (x-axis) labels
                cat_ref = Reference(ws, min_col=min_col, max_col=max_col, 
                                   min_row=1, max_row=1)
                chart.set_categories(cat_ref)
                
            except Exception as e:
                print(f"Error updating chart {chart_idx}: {e}")
    
    def process_patient(self, patient_id, delete_empty_cols=True):
        """Process a single patient and generate Excel output."""
        rows = self.df_main[self.df_main['Screening #'] == patient_id]
        if rows.empty:
            return None
        row = rows.iloc[0]
        
        try:
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
        except Exception as e:
            print(f"Error loading template: {e}")
            return None
        
        # Calculate day offsets from treatment date
        day_data = self.calculate_day_offsets(row)
        num_daily_days = len(day_data)
        
        # Template structure:
        # Col F(6): Screening
        # Col G(7): Baseline (Day -1) - template default
        # Col H(8): Day +1
        # Col I(9): Day +2
        # Col J(10): Day +3 (Discharge)
        # Col K(11): 30D
        # ... etc
        
        # We need to handle dynamic number of daily columns
        # Template has columns G-J for daily (4 columns), but we may need more or fewer
        
        template_daily_cols = 4  # G, H, I, J in template
        
        # Column adjustment is now handled below to account for Discharge column
        
        # Calculate where follow-up columns start (AFTER Discharge)
        # Layout: Screening(F) | Daily(G...) | Discharge | 30D | 6M | ...
        discharge_col = 7 + num_daily_days  # Column right after daily labs
        fu_start_col = discharge_col + 1  # Follow-ups start after Discharge
        
        # Adjust template columns for Discharge
        # Template assumes: G-J=Daily(4), K=30D, L=6M, etc.
        # We need: G...(daily) | Discharge | 30D | ...
        template_cols_before_fu = 4  # G, H, I, J in template = 4 daily cols
        actual_cols_before_fu = num_daily_days + 1  # daily + Discharge
        
        if actual_cols_before_fu > template_cols_before_fu:
            # Insert additional columns
            extra_cols = actual_cols_before_fu - template_cols_before_fu
            insert_at = 11  # Before what was 30D in template
            for i in range(extra_cols):
                ws.insert_cols(insert_at)
                self.copy_column_format(ws, source_col=10, target_col=insert_at)
                ws.cell(row=3, column=insert_at).value = "Result"
        elif actual_cols_before_fu < template_cols_before_fu and delete_empty_cols:
            # Delete excess columns
            cols_to_delete = template_cols_before_fu - actual_cols_before_fu
            for i in range(cols_to_delete):
                col_to_del = 7 + actual_cols_before_fu
                ws.delete_cols(col_to_del)
            # Recalculate positions after deletion
            discharge_col = 7 + num_daily_days
            fu_start_col = discharge_col + 1
        
        # Set daily column headers (row 1) and dates (row 2)
        for idx, (date_str, day_offset, formatted_date) in enumerate(day_data):
            col = 7 + idx  # Start at column G
            is_last = (idx == num_daily_days - 1)
            
            if day_offset is not None:
                # Don't mark as discharge - we have a separate discharge column now
                header = self.format_day_header(day_offset, is_discharge=False)
            else:
                header = f"Day {idx + 1}"
            
            ws.cell(row=1, column=col).value = header
            ws.cell(row=2, column=col).value = formatted_date
        
        # Set Discharge column header and date
        ws.cell(row=1, column=discharge_col).value = "Discharge"
        dv_config = LABS_VISIT_CONFIG.get("Discharge")
        if dv_config and "date_col" in dv_config:
            dv_date = row.get(dv_config["date_col"], "")
            if self.is_valid_value(dv_date):
                ws.cell(row=2, column=discharge_col).value = self.format_date(dv_date)
        
        # Set follow-up column headers and dates
        fu_col = fu_start_col
        for visit_name in FOLLOWUP_VISITS:
            ws.cell(row=1, column=fu_col).value = visit_name
            config = LABS_VISIT_CONFIG.get(visit_name)
            if config and "date_col" in config:
                date_val = row.get(config["date_col"], "")
                if self.is_valid_value(date_val):
                    ws.cell(row=2, column=fu_col).value = self.format_date(date_val)
            fu_col += 1
        
        # Fill in Screening date
        screening_date = row.get("SBV_SV_SVSTDTC", "")
        if self.is_valid_value(screening_date):
            ws.cell(row=2, column=6).value = self.format_date(screening_date)
        
        # Collect reference values for chart parameters (to update Sheet1)
        chart_ref_values = {}
        
        # Fill in data for each parameter row
        # Resolve units first
        self.param_target_units = self.resolve_units(row, num_daily_days, patient_id)

        for template_row, (param_name, test_code_raw, lab_type) in TEMPLATE_ROW_MAP.items():
            # Handle list of test codes (e.g. Troponin)
            if isinstance(test_code_raw, list):
                test_codes = test_code_raw
            else:
                test_codes = [test_code_raw]
            
            # Check which test code has data (for screening/daily columns existence)
            # We use this to decide which column to pull data from
            test_code = None
            for tc in test_codes:
                if self.find_lab_column("SBV_LB_", lab_type, tc):
                    test_code = tc
                    break
            if not test_code:
                test_code = test_codes[0]
            
            # Get resolved target unit for this parameter (row)
            target_unit_info = self.param_target_units.get(template_row)
            
            # Check if this is a dual-unit case
            is_dual = isinstance(target_unit_info, dict) and target_unit_info.get('dual')
            dual_units = target_unit_info.get('units', []) if is_dual else []

            # --- Screening Unit ---
            screening_unit_raw = self.get_units(row, "SBV_LB_", lab_type, test_code)
            # Fallback if missing
            if not screening_unit_raw:
                screening_unit_raw = self.get_units(row, "TV_LB_", lab_type, test_code, pipe_index=0)
            if not screening_unit_raw:
                screening_unit_raw = self.get_units(row, "FU1M_LB_", lab_type, test_code)
            
            screening_unit = self.normalize_unit(screening_unit_raw)

            # Determine Display Unit
            if is_dual:
                # For dual display, show both units in column E
                display_unit = "/".join(dual_units)  # e.g. "fl/%"
            else:
                display_unit = target_unit_info if isinstance(target_unit_info, str) else screening_unit

            # --- Reference Values ---
            ref_min, ref_max, ref_unit = self.get_reference_values(row, lab_type, test_code)
            ref_unit_normalized = self.normalize_unit(ref_unit) if ref_unit else screening_unit
            
            # For dual units, don't convert ref values - just display as-is
            # Convert using the actual ref_unit (from source of ref values), not screening_unit
            if not is_dual and display_unit and ref_unit_normalized and display_unit != ref_unit_normalized:
                ref_min = self.convert_value(ref_min, ref_unit_normalized, display_unit, test_code)
                ref_max = self.convert_value(ref_max, ref_unit_normalized, display_unit, test_code)
            
            if ref_min is not None:
                ws.cell(row=template_row, column=3).value = self.to_number(ref_min)
            if ref_max is not None:
                ws.cell(row=template_row, column=4).value = self.to_number(ref_max)
            
            # --- Set Unit Column (E) ---
            if display_unit:
                unit_cell = ws.cell(row=template_row, column=5)
                if is_dual and len(dual_units) >= 2:
                    # Use rich text to color-code each unit in the combined string
                    from openpyxl.cell.rich_text import TextBlock, CellRichText
                    from openpyxl.cell.text import InlineFont
                    
                    # Create colored text blocks: "unit1" in blue, "/" in black, "unit2" in green
                    blue_font = InlineFont(color=DUAL_UNIT_COLORS['unit1'])
                    green_font = InlineFont(color=DUAL_UNIT_COLORS['unit2'])
                    
                    rich_text = CellRichText(
                        TextBlock(blue_font, dual_units[0]),
                        "/",
                        TextBlock(green_font, dual_units[1])
                    )
                    unit_cell.value = rich_text
                else:
                    unit_cell.value = display_unit
            
            # Collect ref values AND units for chart parameters (Sheet1)
            # WBC=row 25, Cr=row 6, PLT=row 24, CRP=row 41
            chart_param_map = {25: 'WBC', 6: 'Cr', 24: 'PLT', 41: 'CRP'}
            if template_row in chart_param_map:
                chart_ref_values[chart_param_map[template_row]] = (ref_min, ref_max, display_unit)
            
            # --- Screening Value (Column F) ---
            # Special handling for Troponin (row 19) - include T/I label
            if template_row == 19:  # Troponin T/I
                screening_val = self.get_troponin_value(row, "SBV_LB_")
                # Troponin is complex string, skip conversion usually
            else:
                raw_val = self.get_lab_value(row, "SBV_LB_", lab_type, test_code)
                # For dual display, don't convert - just display raw value
                if is_dual:
                    screening_val = raw_val
                else:
                    screening_val = self.convert_value(raw_val, screening_unit, display_unit, test_code)
            
            if screening_val is not None:
                cell = ws.cell(row=template_row, column=6)
                # Don't convert Troponin to number since it has text label
                if template_row == 19:
                    cell.value = screening_val
                else:
                    cell.value = self.to_number(screening_val)
                
                # Apply out-of-range highlighting (red) - takes precedence
                if self.highlight_out_of_range and template_row != 19 and self.is_outside_reference_range(screening_val, ref_min, ref_max):
                    cell.font = Font(color=OUT_OF_RANGE_COLOR)
                # Apply color coding for dual-unit mode
                elif is_dual and len(dual_units) >= 2 and screening_unit:
                    if screening_unit == dual_units[0]:
                        cell.font = Font(color=DUAL_UNIT_COLORS['unit1'])  # Blue
                    elif screening_unit == dual_units[1]:
                        cell.font = Font(color=DUAL_UNIT_COLORS['unit2'])  # Green
            else:
                status = self.get_lab_status(row, "SBV_LB_", lab_type, test_code)
                if status:
                    ws.cell(row=template_row, column=6).value = status
            
            # --- Daily Labs ---
            for day_idx in range(num_daily_days):
                col = 7 + day_idx
                
                # Special handling for Troponin (row 19)
                if template_row == 19:
                    val = self.get_troponin_value(row, "TV_LB_", pipe_index=day_idx)
                    value_unit = None
                else:
                    raw_val = self.get_lab_value(row, "TV_LB_", lab_type, test_code, pipe_index=day_idx)
                    daily_unit_raw = self.get_units(row, "TV_LB_", lab_type, test_code, pipe_index=day_idx)
                    value_unit = self.normalize_unit(daily_unit_raw)
                    
                    # For dual display, don't convert - just display raw value
                    if is_dual:
                        val = raw_val
                    else:
                        d_unit = value_unit if value_unit else display_unit 
                        val = self.convert_value(raw_val, d_unit, display_unit, test_code)
                
                if val is not None:
                    cell = ws.cell(row=template_row, column=col)
                    # Don't convert Troponin to number since it has text label
                    if template_row == 19:
                        cell.value = val
                    else:
                        cell.value = self.to_number(val)
                    
                    # Apply out-of-range highlighting (red) - takes precedence
                    if self.highlight_out_of_range and template_row != 19 and self.is_outside_reference_range(val, ref_min, ref_max):
                        cell.font = Font(color=OUT_OF_RANGE_COLOR)
                    # Apply color coding for dual-unit mode
                    elif is_dual and len(dual_units) >= 2 and value_unit:
                        if value_unit == dual_units[0]:
                            cell.font = Font(color=DUAL_UNIT_COLORS['unit1'])  # Blue
                        elif value_unit == dual_units[1]:
                            cell.font = Font(color=DUAL_UNIT_COLORS['unit2'])  # Green
                else:
                    status = self.get_lab_status(row, "TV_LB_", lab_type, test_code, pipe_index=day_idx)
                    if status:
                        ws.cell(row=template_row, column=col).value = status
            
            # --- Discharge Visit ---
            dv_prefix = LABS_VISIT_CONFIG.get("Discharge", {}).get("prefix", "DV_LB_")
            if template_row == 19:  # Troponin
                dv_val = self.get_troponin_value(row, dv_prefix)
                value_unit = None
            else:
                dv_raw = self.get_lab_value(row, dv_prefix, lab_type, test_code)
                dv_unit_raw = self.get_units(row, dv_prefix, lab_type, test_code)
                value_unit = self.normalize_unit(dv_unit_raw)
                
                if is_dual:
                    dv_val = dv_raw
                else:
                    d_unit = value_unit if value_unit else display_unit
                    dv_val = self.convert_value(dv_raw, d_unit, display_unit, test_code)
            
            if dv_val is not None:
                cell = ws.cell(row=template_row, column=discharge_col)
                if template_row == 19:
                    cell.value = dv_val
                else:
                    cell.value = self.to_number(dv_val)
                
                # Apply out-of-range highlighting (red) - takes precedence
                if self.highlight_out_of_range and template_row != 19 and self.is_outside_reference_range(dv_val, ref_min, ref_max):
                    cell.font = Font(color=OUT_OF_RANGE_COLOR)
                # Apply color coding for dual-unit mode
                elif is_dual and len(dual_units) >= 2 and value_unit:
                    if value_unit == dual_units[0]:
                        cell.font = Font(color=DUAL_UNIT_COLORS['unit1'])
                    elif value_unit == dual_units[1]:
                        cell.font = Font(color=DUAL_UNIT_COLORS['unit2'])
            else:
                status = self.get_lab_status(row, dv_prefix, lab_type, test_code)
                if status:
                    ws.cell(row=template_row, column=discharge_col).value = status
            
            # --- Follow-up Visits ---
            fu_col = fu_start_col
            for visit_name in FOLLOWUP_VISITS:
                config = LABS_VISIT_CONFIG.get(visit_name)
                if not config:
                    fu_col += 1
                    continue
                
                prefix = config["prefix"]
                
                # Special handling for Troponin (row 19)
                if template_row == 19:
                    val = self.get_troponin_value(row, prefix)
                    value_unit = None
                else:
                    raw_val = self.get_lab_value(row, prefix, lab_type, test_code)
                    fu_unit_raw = self.get_units(row, prefix, lab_type, test_code)
                    value_unit = self.normalize_unit(fu_unit_raw)
                    
                    # For dual display, don't convert - just display raw value
                    if is_dual:
                        val = raw_val
                    else:
                        f_unit = value_unit if value_unit else display_unit
                        val = self.convert_value(raw_val, f_unit, display_unit, test_code)
                
                if val is not None:
                    cell = ws.cell(row=template_row, column=fu_col)
                    # Don't convert Troponin to number since it has text label
                    if template_row == 19:
                        cell.value = val
                    else:
                        cell.value = self.to_number(val)
                    
                    # Apply out-of-range highlighting (red) - takes precedence
                    if self.highlight_out_of_range and template_row != 19 and self.is_outside_reference_range(val, ref_min, ref_max):
                        cell.font = Font(color=OUT_OF_RANGE_COLOR)
                    # Apply color coding for dual-unit mode
                    elif is_dual and len(dual_units) >= 2 and value_unit:
                        if value_unit == dual_units[0]:
                            cell.font = Font(color=DUAL_UNIT_COLORS['unit1'])  # Blue
                        elif value_unit == dual_units[1]:
                            cell.font = Font(color=DUAL_UNIT_COLORS['unit2'])  # Green
                else:
                    status = self.get_lab_status(row, prefix, lab_type, test_code)
                    if status:
                        ws.cell(row=template_row, column=fu_col).value = status
                
                fu_col += 1
        
        # Update Sheet1 reference range columns to match daily column count
        self.update_sheet1_columns(wb, num_daily_days, day_data, chart_ref_values)
        
        # Update chart labels (titles and series names)
        self.update_chart_labels(ws, patient_id, chart_ref_values)
        
        # Update charts with peri-procedural data only
        self.update_charts(ws, num_daily_days, day_data)
        
        # Save to BytesIO
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
    
    def generate_export(self, patient_ids, delete_empty_cols=True):
        """Generate export for one or multiple patients."""
        if len(patient_ids) == 1:
            excel_data = self.process_patient(patient_ids[0], delete_empty_cols)
            return (excel_data, 'xlsx', patient_ids[0])
        else:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for pid in patient_ids:
                    excel_data = self.process_patient(pid, delete_empty_cols)
                    if excel_data:
                        zf.writestr(f"{pid}_labs.xlsx", excel_data)
            return (zip_buffer.getvalue(), 'zip', None)
    
    def generate_zip(self, patient_ids, delete_empty_cols=True):
        """Generate ZIP containing Excel files for listed patients."""
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in patient_ids:
                excel_data = self.process_patient(pid, delete_empty_cols)
                if excel_data:
                    zf.writestr(f"{pid}_labs.xlsx", excel_data)
        return zip_buffer.getvalue()
